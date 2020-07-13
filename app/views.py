from django.shortcuts import render
from django.http import HttpResponse
from tablib import Dataset
from datetime import datetime
# from .forms import LekPrice
from django.views.generic import TemplateView, ListView
from django.shortcuts import render

from .models import Leki

import openpyxl

from .resources import LekiResource

from django.db.models import Q


def home_view(request):
    return render(request, 'downloader/base.html')


'''Имена  колонок в базе'''
titul = ['id', 'mnn', 'torgName', 'lekForm', 'factory', 'ATX', 'count', 'predCena',
         'CenaPervUpak', 'ru', 'dateReg', 'EAN13']

###################################################
''' записываем в файл данные'''


def beard_file_csv(data):
    with open('leki.csv', 'a') as leki:
        data = str(data) + ','
        leki.write(data)


######################################################
'''создаем имена будущих колонок в базе'''


def write_titul_in_file(data):
    data = titul

    for name in data:
        beard_file_csv(name)


def read_file(file_name):
    wd = openpyxl.load_workbook(filename=file_name)
    sheet = wd.active

    row_count = sheet.max_row
    column_count = sheet.max_column

    row_count = row_count - (row_count - 1000)
    for row in range(4, row_count):
        beard_file_csv('')
        for col in range(1, column_count - 1):

            work_cell = sheet.cell(row, col).value
            if work_cell == None:
                work_cell = ''
            if type(work_cell) == str:
                work_cell = work_cell.replace(',', '')
                work_cell = '"' + work_cell + '"'
            if type(work_cell) == int:
                work_cell = str(work_cell)
                work_cell = work_cell.replace(',', '.')

            beard_file_csv(work_cell)

    # for cellObj in sheet.columns[:5]:
    #   print(cellObj.value)
    # for row in sheet.rows:
    # for cell in row:
    #     print(cell.value)

    # mnn = sheet['B'+str(3)].value
    # torgName = sheet['B' + str(1)].value
    # print(mnn)


# write_titul_in_file(titul)
# read_file('leki_.xlsx')

def import_leki(request):
    if request.method == 'POST':
        file_format = request.POST['file-format']
        leki_resource = LekiResource()
        dataset = Dataset()
        new_leki = request.FILES['importData']

        if file_format == 'CSV':
            imported_data = dataset.load(new_leki.read().decode('utf-8'), format='csv')
            result = leki_resource.import_data(dataset, dry_run=True)
        elif file_format == 'JSON':
            imported_data = dataset.load(new_leki.read().decode('utf-8'), format='json')
            result = leki_resource.import_data(dataset, dry_run=True)

        if not result.has_errors():
            leki_resource.import_data(dataset, dry_run=False)

    return render(request, 'downloader/import.html')


def export_data(request):
    if request.method == 'POST':
        file_format = request.POST['file-format']
        employee_resource = LekiResource()
        dataset = employee_resource.export()
        if file_format == 'CSV':
            response = HttpResponse(dataset.csv, content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="exported_data.csv"'
            return response
        elif file_format == 'JSON':
            response = HttpResponse(dataset.json, content_type='application/json')
            response['Content-Disposition'] = 'attachment; filename="exported_data.json"'
            return response
        elif file_format == 'XLS (Excel)':
            response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename="exported_data.xls"'
            return response

    return render(request, 'downloader/export.html')


"""
class HomePageView(TemplateView):
    template_name = 'app/index.html'
"""


def get_queryset(request):
    object_list = []
    if request.GET.get('ean13') != None:
        query = request.GET.get('ean13')
        filter_list = Leki.objects.filter(Q(EAN13__icontains=query))
    # date_query = str(request.GET.get('date'))
    #  print(date_query)
    # date_query = datetime.strptime(date_query, "%Y-%m-%d").date()

    # year, month, day = date_query.split('-')
    # date_query = day + '.' + month+ '.' + year
    # print(date_query)
    # date_query = datetime(year, month, day)
    ## print(date_query)




        i = 0
        for r in filter_list:

            lek_dict = {}

            lek_dict['torgName'] = r.torgName
            date_reg = str(r.dateReg)
            date_reg, num_post = date_reg.split()
            date_reg = datetime.strptime(date_reg, "%d.%m.%Y").date()

            # date_reg = datetime.strftime(datetime.strptime(date_reg,'%dd-%BB-%YYYY'),'%Y-%m-%d')
            # if date_query>=date_reg:

            lek_dict['dateReg'] = date_reg
            lek_dict['num_post'] = num_post

            cena = str(r.predCena)

            cena = cena.replace(",", ".")

            cena = float(cena)

            if cena > 500.00:
                nds = round((cena * 0.1), 2)
                nacenka_apt = round((cena * 0.15), 2)
                cena_apt = cena + nacenka_apt + nds
            elif cena > 50.00:
                nds = round((cena * 0.1), 2)

                nacenka_apt = round((cena * 0.28), 2)

                cena_apt = cena + nacenka_apt + nds

            else:
                nds = round((cena * 0.1), 2)
                nacenka_apt = round((cena * 0.32), 2)
                cena_apt = cena + nacenka_apt + nds
            cena_apt = round(cena_apt, 2)
            lek_dict['cena_apt'] = cena_apt
            object_list.append(lek_dict)

    return render(request, "app/index.html",
                      {'object_list': object_list})


